import pytest
from unittest.mock import MagicMock, patch
from src.models.la_search_page import LASearchPage


@pytest.fixture
def mock_browser():
    browser = MagicMock()
    return browser


@pytest.fixture
def search_page(mock_browser):
    la_landing_page = MagicMock()
    la_landing_page.browser = mock_browser
    return LASearchPage(la_landing_page)


def test_get_news(search_page, mock_browser):
    mock_browser.wait_for_element.return_value.find_elements.return_value = [
        MagicMock(
            find_element=MagicMock(
                side_effect=lambda by, selector: MagicMock(text="mocked text")
            ),
            text="mocked text"
        )
    ]

    news_data = search_page.get_news()

    assert len(news_data) > 0
    assert "title" in news_data[0]
    assert "date" in news_data[0]
    assert "description" in news_data[0]
    assert "picture_filename" in news_data[0]
    assert "words-counts_title-description" in news_data[0]
    assert "has_money" in news_data[0]


def test_check_if_subscribe_popup_is_open(search_page, mock_browser):
    mock_browser.wait_for_element_in_shadow.return_value = MagicMock()
    mock_browser.retry_action.return_value = True

    result = search_page.check_if_subscribe_popup_is_open()

    assert result is True

def test_find_sort_button(search_page, mock_browser):
    mock_browser.wait_for_element.return_value = MagicMock()

    search_page.find_sort_button()

    mock_browser.wait_for_element.assert_called()
    mock_browser.wait_for_element.return_value.click.assert_called()


def test_download_image(search_page):
    with patch("requests.get") as mock_get:
        mock_get.return_value.content = b"mocked image content"
        search_page.download_image("http://example.com/image.jpg", "test.jpg")

        with open("output/img/test.jpg", "rb") as file:
            assert file.read() == b"mocked image content"


def test_save_to_excel(search_page):
    news_data = [
        {
            "title": "Sample Title",
            "date": "2024-07-31",
            "description": "Sample description",
            "picture_filename": "1.jpg",
            "words-counts_title-description": 5,
            "has_money": False
        }
    ]

    search_page.save_to_excel(news_data)

    from openpyxl import load_workbook
    wb = load_workbook("output/news_data.xlsx")
    ws = wb.active
    assert ws.title == "News Data"
    assert ws.cell(row=2, column=1).value == "Sample Title"
